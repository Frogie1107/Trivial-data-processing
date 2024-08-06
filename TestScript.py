import os
import pandas as pd
from openpyxl import Workbook,load_workbook

#Check if new sheet was made.If not, new sheet with openpyxl
wb1 = load_workbook('Vehicle information_20240728_213917.xlsx')#copy paste the DMS new generated sheet's name to here

check_sheet = wb1.sheetnames
if 'sorted' in wb1.sheetnames:#check if the file has been created
    print('worksheet already exist')
else:
    ws = wb1.create_sheet(title="sorted")#create a new sheet call 'sorted'
#print(wb1.sheetnames) #check if the worksheet are created
ws = wb1['sorted']


#select the columns from 'Vehicle information' and append to the new sheet
VItable = pd.read_excel('Vehicle information_20240728_213917.xlsx', sheet_name='Vehicle information')#read the excel sheet 'Vehicle information' from excel file

#VIN_column = [3]
#VSname_column = [9]
#Ddate_column = [15]
#country_column = [23] #for reference

copiedColumn = [3, 9, 15, 23] #column of 'VIN','Vehicle series name','Delivery date', 'Target country for vehicle sales' from excel sheet
selected_column = VItable.iloc[:, copiedColumn]
#print(selected_column)
with pd.ExcelWriter('Vehicle information_20240728_213917.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    selected_column.to_excel(writer, sheet_name='sorted', index=False)


#check if the vehicle has delivery date, otherwise remove
ws_sorted = pd.read_excel('Vehicle information_20240728_213917.xlsx', sheet_name='sorted')
#empty_rows = ws_sorted[ws_sorted.iloc[:, 2].isna() | (ws_sorted.iloc[:, 2] == '')]
#print(empty_rows) #test
drop_row = ws_sorted[ws_sorted.iloc[:, 2].isna() | (ws_sorted.iloc[:, 2] == '')]
drop_row = ws_sorted.dropna()
with pd.ExcelWriter('Vehicle information_20240728_213917.xlsx', engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        drop_row.to_excel(writer, sheet_name='sorted', index=False)


wb1.save('Vehicle information_20240728_213917.xlsx')# make changes on the excel file