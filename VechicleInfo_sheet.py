import os
import pandas as pd
from openpyxl import Workbook,load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta


wb_name = 'result1129.xlsx' #the file name to process

result_sheet="Sheet1" #This is the raw data excel sheet name
# result_sheet="Vehicle information" #This is the raw data excel sheet name

wb1 = load_workbook(wb_name)#copy paste the sheet needed to process here 
#Check if new sheet was made.If not, new sheet with openpyxl
check_sheet = wb1.sheetnames
if 'sorted' in wb1.sheetnames:#check if the file has been created
    print('worksheet already exist, processing')
else:
    ws = wb1.create_sheet(title="sorted")#create a new sheet call 'sorted'
    print("new sheet created, processing") 
#print(wb1.sheetnames) 

#select the columns from 'Vehicle information' and append to the new sheet
VItable = pd.read_excel(wb_name, sheet_name=result_sheet)#read the excel sheet 'Vehicle information' from excel file
VItable = pd.read_excel(wb_name, sheet_name=result_sheet)#read the excel sheet 'Vehicle information' from excel file
#VIN_column = [3]
#VSname_column = [10]
#Ddate_column = [16]
#country_column = [24] #for reference
copiedColumn = [3, 10, 16, 24] #column of 'VIN','Vehicle series name','Delivery date', 'Target country for vehicle sales' from excel sheet
selected_column = VItable.iloc[:, copiedColumn]
#print(selected_column)
with pd.ExcelWriter(wb_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    selected_column.to_excel(writer, sheet_name='sorted', index=False)


#check if the vehicle has delivery date, otherwise remove
ws_sorted = pd.read_excel(wb_name, sheet_name='sorted')
ws_sorted = pd.read_excel(wb_name, sheet_name='sorted')
#print(empty_rows) #test
#drop_row = ws_sorted[ws_sorted.iloc[:, 2].isna() | (ws_sorted.iloc[:, 2] == '')]
ws_process = ws_sorted[ws_sorted.iloc[:, 2].isna() | (ws_sorted.iloc[:, 2] == '')]
ws_process = ws_sorted.dropna()#drop_row is the dataframe/worksheet that cleared empty delivery date vehicle
with pd.ExcelWriter(wb_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    ws_process.to_excel(writer, sheet_name='sorted', index=False)


#Replace the country name with abbreviation
Country_code = {
    'Austria':'AT', 
    'Belgium':'BE',
    'Germany':'DE',
    'Denmark':'DK', 
    'Greece': 'GR', 
    'Spain': 'ES',
    'Finland':'FI',
    'France':'FR',
    'United Kingdom':'UK',
    'United Kingdom':'UK',
    'Hungary':'HU',
    'Ireland':'IE',
    'Iceland':'IS',
    'Italy':'IT',
    'Netherlands':'NL',
    'Poland':'PL',
    'Portugal': 'PT',
    'Sweden':'SE'
}
#country_code = ws_process
ws_process.iloc[:,3] = ws_process.iloc[:,3].replace(Country_code)
with pd.ExcelWriter(wb_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
   ws_process.to_excel(writer, sheet_name='sorted', index=False)


#Replace the model name ######EXPANDING
model_code = {
    'Tang EV':'TANG',#TANG
    'EU TANG EV LHD 2023':'TANG',
    'HAN EV GB/T':'HAN',#HAN
    'BYD ATTO 3 LHD':'ATTO 3' ,#ATTO3
    'BYD ATTO 3 RHD': 'ATTO 3',
    '2024 BYD ATTO 3 LHD':'ATTO 3',
    'Dolphin LHD': 'DOLPHIN',#DOLPHIN
    'Dolphin RHD':'DOLPHIN',
    'Seal LHD':'SEAL', #SEAL
    'Seal RHD':'SEAL',
    'ETP3 LHD':'T3',#T3
    'ETP3 RHD':'T3',
    'EU SONG PLUS EV LHD 2023':'SEAL U EV', #SEAL U
    'EU SONG PLUS DMI LHD 2023':'SEAL U DM-i',
    'EU SONG PLUS DMI LHD 2023':'SEAL U DM-i',
    'EU SONG PLUS DMI RHD 2023':'SEAL U DM-i',
    'EU SEALION 7 EV LHD 2024':'SEALION 7',
    'UZ_SONGPLUS_DMI_LHD_2023':'',#invalid name for EU
    'UZ_SONGPRO_DMI_LHD_2023':'',
    'Chaser UZ':'',
    'Han EV UZ':'',
    'Song Plus UZ':'',
    'Song Plus EV UZ 2023':'',
}
#model_convert = ws_process
ws_process.iloc[:,1] = ws_process.iloc[:,1].replace(model_code)
with pd.ExcelWriter(wb_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
    ws_process.to_excel(writer, sheet_name='sorted', index=False)


#move the columns and reorder them
wb2 = load_workbook(wb_name)# B U G !!! I have to load the workbook again so that the worksheet can be updated
ws_move_col = wb2['sorted']
max_row=ws_move_col.max_row # max_row = 1 in the first run
column_mapping = {
    4:7,
    3:5,
    2:4,
    1:2,
}
#print(max_row)
for row in range(1,max_row+1):
    for src_col, dest_col in column_mapping.items():
        cell_value = ws_move_col.cell(row=row, column=src_col).value
        ws_move_col.cell(row=row, column=dest_col).value = cell_value
        # Optionally clear the original cell
        ws_move_col.cell(row=row, column=src_col).value = None
for row in range(2,max_row+1):
    ws_move_col.cell(row=row, column=3).value = 'BYD' 
    ws_move_col.cell(row=row, column=5).value = datetime.strptime(ws_move_col.cell(row=row, column=5).value,"%Y-%m-%d")
    ws_move_col.cell(row=row, column=6).value = ws_move_col.cell(row=row, column=5).value + relativedelta(years=2)
    ws_move_col.cell(row=row, column=5).value = ws_move_col.cell(row=row, column=5).value.strftime("%d/%m/%Y") #Convert to UK time format
    ws_move_col.cell(row=row, column=6).value = ws_move_col.cell(row=row, column=6).value.strftime("%d/%m/%Y") #Convert to UK time format
# change the column title
col_title = ['LICENSE PLATE','VIN','MAKE','MODEL','COVERAGE VALID FROM','COVERAGE VALID TO','COUNTRY']
for col_index, value in enumerate(col_title, start=1):
    cell = ws_move_col.cell(row=1, column=col_index)
    cell.value = value


wb2.save('sorted.xlsx')# make changes on the excel file
df = pd.read_excel('sorted.xlsx', sheet_name='sorted', engine='openpyxl')
df.to_csv('sorted.csv', index=False)
print('Process Done!')
#ws_process.to_excel("Sorted.xlsx", index=False)#save the sorted sheet onto a new excel.